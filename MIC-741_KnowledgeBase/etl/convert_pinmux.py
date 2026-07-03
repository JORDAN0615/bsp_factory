"""Convert material/*.xlsm pinmux templates → corpus/pinmux/*.md

Each pin becomes a Markdown section:

    ## SYS_RESET_N
    - Pin: L60 | MPIO: SF_RST_N
    - Direction: I
    - Function: RESET (SFIO0)
    - Pad: sys_reset_n | Domain: vddio_sys_ao
    - POR: JT_RST | Pull: 20K

Usage:
    python etl/convert_pinmux.py [--force]
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

MATERIAL_DIR = Path(__file__).parent.parent / "material"
CORPUS_DIR   = Path(__file__).parent.parent / "corpus" / "pinmux"

# Column indices in the pinmux sheet (0-based)
# Derived by inspecting Row 8 of the IGX_T5000 / Jetson Thor sheets
_COL = {
    "pin_num":      0,
    "signal_name":  1,
    "mpio":         2,
    "verilog_name": 3,
    "sfio0":        6,   # Primary alternate function
    "sfio1":        7,
    "sfio2":        8,
    "sfio3":        9,
    "dir_gpio":     10,
    "dir_sfio0":    11,
    "pad_name":     20,
    "voltage_dom":  21,
    "por":          27,
    "pull":         29,
    "customer":     33,
}

# Row where actual pin data starts (1-based, as openpyxl returns)
_DATA_START_ROW = 10


def _cell(row: tuple, col: int) -> str:
    """Safe cell read: return empty string if out of range or None."""
    if col < len(row) and row[col] is not None:
        return str(row[col]).strip()
    return ""


def _is_group_header(row: tuple) -> bool:
    """Rows with signal in col 2 but no pin number are voltage-group headers."""
    return bool(_cell(row, 2)) and not _cell(row, 0) and not _cell(row, 1)


def _convert_sheet(ws, sheet_name: str, source_name: str) -> str:
    """Convert one pinmux sheet to Markdown text."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < _DATA_START_ROW:
        return ""

    lines: list[str] = [f"# {source_name} — {sheet_name} Pinmux\n"]
    current_group = ""

    for row in rows[_DATA_START_ROW - 1:]:   # 0-based index
        if all(c is None for c in row):
            continue

        # Voltage group header (e.g. VDDIO_A)
        if _is_group_header(row):
            group = _cell(row, 2)
            if group and group != current_group:
                current_group = group
                lines.append(f"\n---\n### Group: {group}\n")
            continue

        pin  = _cell(row, _COL["pin_num"])
        sig  = _cell(row, _COL["signal_name"])
        mpio = _cell(row, _COL["mpio"])

        if not sig and not mpio:
            continue

        label = sig or mpio
        lines.append(f"\n## {label}")

        details: list[str] = []
        if pin:
            details.append(f"Pin: {pin}")
        if mpio and mpio != label:
            details.append(f"MPIO: {mpio}")
        vname = _cell(row, _COL["verilog_name"])
        if vname and vname != label:
            details.append(f"Verilog: {vname}")
        if details:
            lines.append("- " + " | ".join(details))

        # Functions
        funcs = [_cell(row, _COL[f"sfio{i}"]) for i in range(4)]
        funcs = [f for f in funcs if f]
        if funcs:
            lines.append(f"- Function: {' / '.join(funcs)}")

        # Direction
        dir_g = _cell(row, _COL["dir_gpio"])
        dir_s = _cell(row, _COL["dir_sfio0"])
        dirs = [d for d in [dir_g, dir_s] if d]
        if dirs:
            lines.append(f"- Direction: {' / '.join(set(dirs))}")

        # Pad
        pad = _cell(row, _COL["pad_name"])
        dom = _cell(row, _COL["voltage_dom"])
        pad_parts = [p for p in [pad, dom] if p]
        if pad_parts:
            lines.append(f"- Pad: {' | '.join(pad_parts)}")

        # POR + Pull
        por  = _cell(row, _COL["por"])
        pull = _cell(row, _COL["pull"])
        misc = [p for p in [f"POR: {por}" if por else "", f"Pull: {pull}" if pull else ""] if p]
        if misc:
            lines.append(f"- {' | '.join(misc)}")

        # Customer config
        cust = _cell(row, _COL["customer"])
        if cust:
            lines.append(f"- Customer: {cust}")

    return "\n".join(lines)


def convert_xlsm(xlsm_path: Path, output_dir: Path, force: bool = False) -> list[Path]:
    try:
        import openpyxl
    except ImportError:
        print("[ERROR] openpyxl not installed. Run: uv add openpyxl")
        sys.exit(1)

    output_dir.mkdir(parents=True, exist_ok=True)
    outputs: list[Path] = []

    wb = openpyxl.load_workbook(xlsm_path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        # Skip readme/guide sheets
        if any(k in sheet_name.lower() for k in ("readme", "quick guide", "guide")):
            continue

        out_name = f"{xlsm_path.stem}_{sheet_name.replace(' ', '_')}.md"
        out_path = output_dir / out_name

        if out_path.exists() and not force:
            print(f"  [SKIP] {out_name} (already converted, use --force to redo)")
            outputs.append(out_path)
            continue

        ws = wb[sheet_name]
        content = _convert_sheet(ws, sheet_name, xlsm_path.stem)
        if not content.strip():
            continue

        out_path.write_text(content, encoding="utf-8")
        pin_count = content.count("\n## ")
        size_kb = out_path.stat().st_size // 1024
        print(f"  [OK] {xlsm_path.name}/{sheet_name} → {out_name} ({pin_count} pins, {size_kb} KB)")
        outputs.append(out_path)

    wb.close()
    return outputs


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert pinmux Excel files to Markdown")
    parser.add_argument("--force", action="store_true", help="Re-convert even if output exists")
    args = parser.parse_args()

    files = sorted(MATERIAL_DIR.glob("*.xls*"))
    if not files:
        print(f"No Excel files found in {MATERIAL_DIR}")
        return

    print(f"Converting {len(files)} Excel file(s) → {CORPUS_DIR}/")
    for f in files:
        convert_xlsm(f, CORPUS_DIR, force=args.force)
    print("Done.")


if __name__ == "__main__":
    main()
