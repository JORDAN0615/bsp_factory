import os
import re
from pathlib import Path

from rag.models import Chunk

# KB directory — override with KB_DIR env var
_DEFAULT_KB_DIR = "MIC-741_KnowledgeBase"
KB_DIR = Path(os.getenv("KB_DIR", _DEFAULT_KB_DIR))

EXCLUDED_FILES = {"RE-INDEX.md"}

# Max pages per PDF chunk (3 pages ≈ 800-1200 words)
_PDF_PAGES_PER_CHUNK = 3


# ── Markdown helpers ──────────────────────────────────────────────────────────

def _split_by_sections(text: str, source: str) -> list[Chunk]:
    """Split by ## headings, keeping heading as section label."""
    chunks: list[Chunk] = []
    parts = re.split(r"(?=^## )", text, flags=re.MULTILINE)
    for i, part in enumerate(parts):
        part = part.strip()
        if not part:
            continue
        lines = part.splitlines()
        section = lines[0].lstrip("# ").strip() if lines else ""
        chunks.append(Chunk(id=f"{source}::{i}", text=part, source=source, section=section))
    return chunks


def _load_glob_md(kb_dir: Path, pattern: str) -> list[Chunk]:
    chunks: list[Chunk] = []
    for path in sorted(kb_dir.glob(pattern)):
        if path.name in EXCLUDED_FILES:
            continue
        try:
            text = path.read_text(encoding="utf-8")
        except (UnicodeDecodeError, OSError):
            continue
        source = path.relative_to(kb_dir.parent).as_posix()
        chunks.extend(_split_by_sections(text, source))
    return chunks


# ── Patch loader ──────────────────────────────────────────────────────────────

def _split_patch(text: str, source: str) -> list[Chunk]:
    """Split a .patch file into per-file diff chunks."""
    chunks: list[Chunk] = []
    # Split by 'diff --git' markers; first part is the commit header
    parts = re.split(r"(?=^diff --git )", text, flags=re.MULTILINE)
    for i, part in enumerate(parts):
        part = part.strip()
        if not part or len(part) < 30:
            continue
        # Section label: first non-empty line
        section = part.splitlines()[0][:120]
        chunks.append(Chunk(id=f"{source}::{i}", text=part, source=source, section=section))
    return chunks


def load_patch_chunks(kb_dir: Path = KB_DIR) -> list[Chunk]:
    """Load .patch files from 02_Original_Code and 03_Git_History/patches."""
    chunks: list[Chunk] = []
    patterns = [
        "02_Original_Code/**/*.patch",
        "02_Original_Code/**/*.diff",
        "03_Git_History/patches/*.patch",
    ]
    for pattern in patterns:
        for path in sorted(kb_dir.glob(pattern)):
            try:
                text = path.read_text(encoding="utf-8", errors="replace")
            except OSError:
                continue
            source = path.relative_to(kb_dir.parent).as_posix()
            chunks.extend(_split_patch(text, source))
    return chunks


# ── PDF loader ────────────────────────────────────────────────────────────────

def load_pdf_chunks(kb_dir: Path = KB_DIR) -> list[Chunk]:
    """Extract text from material/*.pdf and chunk by page groups."""
    chunks: list[Chunk] = []
    pdf_dir = kb_dir / "material"
    if not pdf_dir.exists():
        return chunks

    try:
        import pdfplumber
    except ImportError:
        print("[WARN] pdfplumber not installed — skipping PDF indexing. Run: uv add pdfplumber")
        return chunks

    for pdf_path in sorted(pdf_dir.glob("*.pdf")):
        source = pdf_path.relative_to(kb_dir.parent).as_posix()
        try:
            with pdfplumber.open(pdf_path) as pdf:
                pages = pdf.pages
                total = len(pages)
                for start in range(0, total, _PDF_PAGES_PER_CHUNK):
                    group = pages[start: start + _PDF_PAGES_PER_CHUNK]
                    texts = []
                    for page in group:
                        t = page.extract_text()
                        if t:
                            texts.append(t.strip())
                    combined = "\n\n".join(texts)
                    if len(combined) < 50:
                        continue
                    section = f"p{start+1}-{min(start+_PDF_PAGES_PER_CHUNK, total)}"
                    chunks.append(Chunk(
                        id=f"{source}::{start}",
                        text=combined,
                        source=source,
                        section=section,
                    ))
        except Exception as e:
            print(f"[WARN] PDF parse failed ({pdf_path.name}): {e}")

    return chunks


# ── Excel loader ─────────────────────────────────────────────────────────────

_EXCEL_ROWS_PER_CHUNK = 50   # rows per chunk (pinmux sheets can be 300+ rows)


def load_excel_chunks(kb_dir: Path = KB_DIR) -> list[Chunk]:
    """Extract text from material/*.xlsx and *.xlsm (pinmux templates, BOM, etc.)."""
    chunks: list[Chunk] = []
    pdf_dir = kb_dir / "material"
    if not pdf_dir.exists():
        return chunks

    try:
        import openpyxl
    except ImportError:
        print("[WARN] openpyxl not installed — skipping Excel indexing. Run: uv add openpyxl")
        return chunks

    for xls_path in sorted(pdf_dir.glob("*.xls*")):
        source = xls_path.relative_to(kb_dir.parent).as_posix()
        try:
            wb = openpyxl.load_workbook(xls_path, read_only=True, data_only=True)
        except Exception as e:
            print(f"[WARN] Excel open failed ({xls_path.name}): {e}")
            continue

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # First non-empty row as header
            header = [str(c).strip() if c is not None else "" for c in rows[0]]
            data_rows = rows[1:]

            chunk_idx = 0
            for start in range(0, len(data_rows), _EXCEL_ROWS_PER_CHUNK):
                batch = data_rows[start: start + _EXCEL_ROWS_PER_CHUNK]
                lines: list[str] = [f"[{xls_path.stem} / {sheet_name}]"]
                for row in batch:
                    # Pair header with value, skip all-empty rows
                    pairs = [
                        f"{h}={v}" for h, v in zip(header, row)
                        if h and v is not None and str(v).strip()
                    ]
                    if pairs:
                        lines.append(" | ".join(pairs))

                text = "\n".join(lines)
                if len(text) < 80:
                    continue

                section = f"{sheet_name} rows {start+1}-{min(start+_EXCEL_ROWS_PER_CHUNK, len(data_rows))}"
                chunks.append(Chunk(
                    id=f"{source}::{sheet_name}::{chunk_idx}",
                    text=text,
                    source=source,
                    section=section,
                ))
                chunk_idx += 1

        wb.close()

    return chunks


# ── Corpus loader (ETL-converted Markdown) ───────────────────────────────────

_CORPUS_LINES_PER_CHUNK = 100   # ~400-600 words per chunk


def _load_corpus(corpus_dir: Path) -> list[Chunk]:
    """Load corpus/ Markdown with file-type aware chunking.

    - design-guide/ (PDF-derived): fixed 100-line chunks — PDFs have too many
      headings to split by ## without exploding chunk count.
    - pinmux/ (Excel-derived): split by ## (each pin is already one section).
    """
    chunks: list[Chunk] = []
    for path in sorted(corpus_dir.rglob("*.md")):
        try:
            text = path.read_text(encoding="utf-8")
        except OSError:
            continue
        source = path.relative_to(corpus_dir.parent).as_posix()

        # pinmux files already have one ## per pin — keep heading-based split
        if "pinmux" in path.parts:
            chunks.extend(_split_by_sections(text, source))
        else:
            # PDF-derived: fixed-line chunks to avoid heading explosion
            chunks.extend(_split_fixed(text, source, lines_per_chunk=_CORPUS_LINES_PER_CHUNK))

    return chunks


# ── Source code loader ───────────────────────────────────────────────────────

_SOURCE_EXTENSIONS = {'.c', '.h', '.cpp', '.dtsi', '.dts', '.sh', '.mk', '.yaml', '.yml'}
_MIN_CHUNK_LEN = 80


def _split_by_blocks(text: str, source: str) -> list[Chunk]:
    """Split by top-level brace blocks {…} — works for C, DTSI, shell.

    Tracks brace depth; each complete top-level block becomes one chunk.
    Preamble (before first block) is kept as its own chunk.
    """
    chunks: list[Chunk] = []
    lines  = text.splitlines()
    depth  = 0
    start  = 0

    for i, line in enumerate(lines):
        # Ignore braces inside comments (best-effort)
        stripped = re.sub(r'//.*', '', line)
        stripped = re.sub(r'/\*.*?\*/', '', stripped)
        opens  = stripped.count('{')
        closes = stripped.count('}')

        if depth == 0 and opens > 0:
            # Flush preamble before this block
            pre = '\n'.join(lines[start:i]).strip()
            if len(pre) >= _MIN_CHUNK_LEN:
                chunks.append(Chunk(
                    id=f"{source}::{start}", text=pre,
                    source=source, section=lines[start][:80].strip(),
                ))
            start = i

        depth += opens - closes

        if depth <= 0 and i >= start:
            block = '\n'.join(lines[start:i + 1]).strip()
            if len(block) >= _MIN_CHUNK_LEN:
                section = lines[start][:80].strip()
                chunks.append(Chunk(
                    id=f"{source}::{start}", text=block,
                    source=source, section=section,
                ))
            start  = i + 1
            depth  = 0

    # Tail
    tail = '\n'.join(lines[start:]).strip()
    if len(tail) >= _MIN_CHUNK_LEN:
        chunks.append(Chunk(
            id=f"{source}::tail", text=tail,
            source=source, section="(tail)",
        ))

    return chunks


def _split_shell(text: str, source: str) -> list[Chunk]:
    """Split shell scripts by function boundaries."""
    # Match:  func_name() {  or  function func_name {
    pattern = re.compile(r'^(?:function\s+\w+|\w+\s*\(\s*\))\s*\{', re.MULTILINE)
    positions = [m.start() for m in pattern.finditer(text)]
    if not positions:
        return _split_by_blocks(text, source)

    chunks: list[Chunk] = []
    boundaries = positions + [len(text)]
    # Preamble
    pre = text[:positions[0]].strip()
    if len(pre) >= _MIN_CHUNK_LEN:
        chunks.append(Chunk(id=f"{source}::0", text=pre, source=source, section="preamble"))

    for idx, (s, e) in enumerate(zip(boundaries, boundaries[1:])):
        block = text[s:e].strip()
        if len(block) >= _MIN_CHUNK_LEN:
            section = block.splitlines()[0][:80]
            chunks.append(Chunk(id=f"{source}::{s}", text=block, source=source, section=section))

    return chunks


def _split_fixed(text: str, source: str, lines_per_chunk: int = 60) -> list[Chunk]:
    """Fallback: split by fixed line count."""
    lines  = text.splitlines()
    chunks: list[Chunk] = []
    for start in range(0, len(lines), lines_per_chunk):
        block = '\n'.join(lines[start:start + lines_per_chunk]).strip()
        if len(block) >= _MIN_CHUNK_LEN:
            section = lines[start][:80].strip() if lines[start:] else ""
            chunks.append(Chunk(id=f"{source}::{start}", text=block, source=source, section=section))
    return chunks


def load_source_chunks(kb_dir: Path = KB_DIR) -> list[Chunk]:
    """Load before/ and after/ source code from 02_Original_Code/, split by language."""
    chunks: list[Chunk] = []
    for variant in ('before', 'after'):
        for path in sorted(kb_dir.glob(f"02_Original_Code/*/{variant}/**/*")):
            if not path.is_file() or path.suffix.lower() not in _SOURCE_EXTENSIONS:
                continue
            try:
                text = path.read_text(encoding='utf-8', errors='replace')
            except OSError:
                continue
            if len(text) < _MIN_CHUNK_LEN:
                continue

            source = path.relative_to(kb_dir.parent).as_posix()
            ext    = path.suffix.lower()

            if ext in ('.c', '.h', '.cpp', '.dtsi', '.dts'):
                file_chunks = _split_by_blocks(text, source)
            elif ext == '.sh':
                file_chunks = _split_shell(text, source)
            else:
                file_chunks = _split_fixed(text, source)

            # Prefix each chunk text with file path for BM25 keyword matching
            issue_dir = path.parts[-4] if len(path.parts) >= 4 else ""
            header    = f"# [{variant}] {issue_dir} — {path.name}\n"
            for c in file_chunks:
                chunks.append(Chunk(
                    id=c.id, text=header + c.text,
                    source=c.source, section=c.section,
                ))

    return chunks


# ── Main entry point ──────────────────────────────────────────────────────────

def load_chunks(kb_dir: Path = KB_DIR) -> list[Chunk]:
    """Load all indexable KB documents into Qdrant chunks."""
    if not kb_dir.exists():
        print(f"[WARN] KB_DIR not found: {kb_dir}")
        return []

    chunks: list[Chunk] = []

    md_chunks = (
        _load_glob_md(kb_dir, "01_Issues/*.md")
        + _load_glob_md(kb_dir, "02_Original_Code/*/PROMPT.md")
        + _load_glob_md(kb_dir, "03_Git_History/*.md")
    )
    patch_chunks = load_patch_chunks(kb_dir)

    # corpus/ = pre-converted Markdown from ETL scripts (preferred over raw PDF/Excel)
    corpus_dir = kb_dir / "corpus"
    if corpus_dir.exists():
        corpus_chunks = _load_corpus(corpus_dir)
        pdf_chunks    = []
        excel_chunks  = []
        print(f"  [corpus] Using pre-converted Markdown from corpus/ ({len(corpus_chunks)} chunks)")
    else:
        corpus_chunks = []
        pdf_chunks    = load_pdf_chunks(kb_dir)
        excel_chunks  = load_excel_chunks(kb_dir)

    source_chunks = load_source_chunks(kb_dir)

    chunks = md_chunks + patch_chunks + source_chunks + corpus_chunks + pdf_chunks + excel_chunks

    print(f"  Loaded {len(chunks)} chunks total "
          f"(md={len(md_chunks)}, patch={len(patch_chunks)}, "
          f"source={len(source_chunks)}, corpus={len(corpus_chunks)}, "
          f"pdf={len(pdf_chunks)}, excel={len(excel_chunks)})")
    return chunks
