from __future__ import annotations

import re
import zipfile
from datetime import date, datetime, time
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from agent.tools.doc_ingest.manifest import ManifestSource
from agent.tools.doc_ingest.records import IngestedRecords, PinRecord
from agent.tools.mic741_knowledge import KnowledgeDBError, _extract_symbols


_BALL_CANDIDATES = ("ball", "ball name")
_SIGNAL_CANDIDATES = ("signal", "signal name")
_XML_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_XML_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_XML_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"


def ingest_pinmux(source: ManifestSource) -> IngestedRecords:
    try:
        import openpyxl
    except ImportError as exc:
        raise KnowledgeDBError("openpyxl is required for pinmux document ingestion") from exc
    if source.header_rows is None:
        raise KnowledgeDBError(f"pinmux source has no header_rows: {source.doc_key}")
    workbook = openpyxl.load_workbook(source.path, data_only=True, read_only=True)
    try:
        worksheet = _select_sheet(workbook, source)
        headers = flatten_headers(
            worksheet,
            source.header_rows,
            merged_ranges=_merged_ranges(source.path, worksheet.title),
        )
        pins = extract_pin_records(worksheet, headers, source)
        return IngestedRecords(pins=pins)
    finally:
        workbook.close()


def flatten_headers(
    worksheet: Any,
    header_rows: tuple[int, int],
    *,
    merged_ranges: list[tuple[int, int, int, int]] | None = None,
) -> list[str]:
    start, end = header_rows
    rows = [list(row) for row in worksheet.iter_rows(min_row=start, max_row=end, values_only=True)]
    width = max([worksheet.max_column or 0, *(len(row) for row in rows)], default=0)
    matrix = [row + [None] * (width - len(row)) for row in rows]
    for min_col, min_row, max_col, max_row in merged_ranges or []:
        if max_row < start or min_row > end or min_col > width:
            continue
        top_row = min_row - start
        if not 0 <= top_row < len(matrix):
            continue
        value = matrix[top_row][min_col - 1]
        if _is_empty(value):
            continue
        for row_no in range(max(start, min_row), min(end, max_row) + 1):
            for col_no in range(min_col, min(max_col, width) + 1):
                matrix[row_no - start][col_no - 1] = value

    raw_names: list[str] = []
    for column in range(width):
        parts: list[str] = []
        for row in matrix:
            value = _clean_cell(row[column])
            if value and (not parts or parts[-1] != value):
                parts.append(value)
        raw_names.append(" ".join(parts) if parts else f"column_{column + 1}")
    return _deduplicate(raw_names)


def extract_pin_records(
    worksheet: Any,
    headers: list[str],
    source: ManifestSource,
) -> list[PinRecord]:
    ball_index = _find_header(headers, _BALL_CANDIDATES)
    signal_index = _find_header(headers, _SIGNAL_CANDIDATES)
    records: list[PinRecord] = []
    start_row = (source.header_rows or (1, 1))[1] + 1
    for values in worksheet.iter_rows(min_row=start_row, max_col=len(headers), values_only=True):
        if all(_is_empty(value) for value in values):
            continue
        columns = {
            headers[index]: _json_value(value)
            for index, value in enumerate(values)
            if index < len(headers) and not _is_empty(value)
        }
        if not columns:
            continue
        ball = _optional_text(values[ball_index]) if ball_index is not None else None
        signal = _optional_text(values[signal_index]) if signal_index is not None else None
        if (ball_index is not None or signal_index is not None) and not (ball or signal):
            continue
        details = "; ".join(f"{name}={value}" for name, value in columns.items())
        identity = " ".join(value for value in [ball, signal] if value) or "(unidentified)"
        content = (
            f"Pin {identity} — {source.platform}, pinmux template v{source.version}. "
            f"{details}"
        )
        records.append(
            PinRecord(
                platform=source.platform,
                ball=ball,
                signal_name=signal,
                columns=columns,
                content=content,
                symbols=_extract_symbols(content),
            )
        )
    return records


def _select_sheet(workbook: Any, source: ManifestSource) -> Any:
    if source.sheet:
        if source.sheet not in workbook.sheetnames:
            raise KnowledgeDBError(
                f"pinmux sheet {source.sheet!r} not found in {source.source_path}"
            )
        return workbook[source.sheet]
    assert source.header_rows is not None
    candidates: list[tuple[int, str, Any]] = []
    for worksheet in workbook.worksheets:
        headers = flatten_headers(
            worksheet,
            source.header_rows,
            merged_ranges=_merged_ranges(source.path, worksheet.title),
        )
        score = int(_find_header(headers, _BALL_CANDIDATES) is not None)
        score += int(_find_header(headers, _SIGNAL_CANDIDATES) is not None)
        if score:
            candidates.append((score, worksheet.title, worksheet))
    if not candidates:
        raise KnowledgeDBError(
            f"could not auto-detect a pin sheet in {source.source_path}; "
            "set sheet in the manifest"
        )
    return sorted(candidates, key=lambda item: (-item[0], item[1]))[0][2]


def _find_header(headers: list[str], candidates: tuple[str, ...]) -> int | None:
    normalized_candidates = [candidate.split() for candidate in candidates]
    best: tuple[int, int] | None = None
    for index, header in enumerate(headers):
        words = re.findall(r"[a-z0-9]+", header.lower())
        for candidate in normalized_candidates:
            if words == candidate:
                score = 3
            elif len(words) >= len(candidate) and words[-len(candidate) :] == candidate:
                score = 2
            elif all(word in words for word in candidate):
                score = 1
            else:
                continue
            if best is None or score > best[0]:
                best = (score, index)
    return best[1] if best else None


def _deduplicate(names: list[str]) -> list[str]:
    counts: dict[str, int] = {}
    result: list[str] = []
    for name in names:
        count = counts.get(name, 0) + 1
        counts[name] = count
        result.append(name if count == 1 else f"{name}_{count}")
    return result


def _clean_cell(value: Any) -> str:
    return " ".join(str(value).split()) if not _is_empty(value) else ""


def _is_empty(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _optional_text(value: Any) -> str | None:
    text = _clean_cell(value)
    return text or None


def _json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _merged_ranges(path: Path, sheet_name: str) -> list[tuple[int, int, int, int]]:
    from openpyxl.utils.cell import range_boundaries

    try:
        with zipfile.ZipFile(path) as archive:
            workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
            relationships = ElementTree.fromstring(
                archive.read("xl/_rels/workbook.xml.rels")
            )
            relation_targets = {
                relation.attrib["Id"]: relation.attrib["Target"]
                for relation in relationships.findall(f"{{{_XML_PKG_REL}}}Relationship")
            }
            sheet_target = None
            for sheet in workbook.findall(f".//{{{_XML_MAIN}}}sheet"):
                if sheet.attrib.get("name") == sheet_name:
                    sheet_target = relation_targets.get(sheet.attrib.get(f"{{{_XML_REL}}}id", ""))
                    break
            if not sheet_target:
                return []
            target = sheet_target.lstrip("/")
            if not target.startswith("xl/"):
                target = f"xl/{target}"
            worksheet = ElementTree.fromstring(archive.read(target))
            return [
                range_boundaries(cell.attrib["ref"])
                for cell in worksheet.findall(f".//{{{_XML_MAIN}}}mergeCell")
            ]
    except (KeyError, OSError, zipfile.BadZipFile, ElementTree.ParseError):
        return []
