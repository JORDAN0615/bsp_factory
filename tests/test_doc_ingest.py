from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
import yaml

from agent.tools.doc_ingest.manifest import ManifestSource, load_manifest
from agent.tools.doc_ingest.pdf_docs import (
    _MIN_PROSE_CHARS,
    _MIN_TABLE_CHARS,
    _emit_figure_ref,
    _emit_prose_chunks,
    _is_heading,
    _render_table,
    _strip_boilerplate,
    chunk_prose,
    detect_boilerplate,
    ingest_pdf,
)
from agent.tools.doc_ingest.pinmux_xlsm import (
    _merged_ranges,
    flatten_headers,
    ingest_pinmux,
)
from agent.tools.doc_ingest.records import IngestedRecords, _content_hash, _replace_source
from agent.tools.mic741_knowledge import KnowledgeDBError


def _source(path: Path, *, header_rows: tuple[int, int] = (1, 2)) -> ManifestSource:
    return ManifestSource(
        path=path,
        source_path=str(path),
        doc_key="test-pinmux",
        title="Test Pinmux",
        doc_type="pinmux_template",
        platform="Thor",
        version="1.0",
        sheet="Pins",
        header_rows=header_rows,
    )


def _save_workbook(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Pins"
    for column, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=column, value=header)
    for row_index, values in enumerate(rows, start=2):
        for column, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column, value=value)
    workbook.save(path)
    workbook.close()


def test_flatten_headers_propagates_merged_cells_and_deduplicates(tmp_path: Path) -> None:
    path = tmp_path / "merged.xlsm"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Pins"
    sheet["A1"] = "Pin Identity"
    sheet.merge_cells("A1:B1")
    sheet["A2"] = "Ball Name"
    sheet["B2"] = "Signal Name"
    sheet["C1"] = "Mode"
    sheet["D1"] = "Mode"
    workbook.save(path)
    workbook.close()

    loaded = openpyxl.load_workbook(path, data_only=True, read_only=True)
    headers = flatten_headers(
        loaded["Pins"],
        (1, 2),
        merged_ranges=_merged_ranges(path, "Pins"),
    )
    loaded.close()

    assert headers == [
        "Pin Identity Ball Name",
        "Pin Identity Signal Name",
        "Mode",
        "Mode_2",
    ]


def test_pinmux_emits_one_record_per_pin_with_every_populated_column(tmp_path: Path) -> None:
    path = tmp_path / "pins.xlsm"
    _save_workbook(
        path,
        ["Ball Name", "Signal Name", "Direction", "Mux Option"],
        [["A12", "I2C3_SCL", "input", "i2c3"], ["B13", "CAN2_TX", "output", "mttcan2"]],
    )

    records = ingest_pinmux(_source(path, header_rows=(1, 1))).pins

    assert len(records) == 2
    assert records[0].ball == "A12"
    assert records[0].signal_name == "I2C3_SCL"
    assert records[0].columns == {
        "Ball Name": "A12",
        "Signal Name": "I2C3_SCL",
        "Direction": "input",
        "Mux Option": "i2c3",
    }
    assert "I2C3_SCL" in records[0].content


def test_pinmux_leaves_ball_and_signal_null_when_columns_are_absent(tmp_path: Path) -> None:
    path = tmp_path / "no-identity.xlsm"
    _save_workbook(path, ["Index", "Direction"], [[1, "input"]])

    records = ingest_pinmux(_source(path, header_rows=(1, 1))).pins

    assert len(records) == 1
    assert records[0].ball is None
    assert records[0].signal_name is None
    assert records[0].columns == {"Index": 1, "Direction": "input"}


def test_boilerplate_requires_more_than_half_the_pages() -> None:
    boilerplate = detect_boilerplate(
        [
            ["NVIDIA CONFIDENTIAL", "Only page one"],
            ["NVIDIA CONFIDENTIAL", "Only page two"],
            ["NVIDIA CONFIDENTIAL", "Unique"],
            ["Different", "Unique four"],
        ]
    )

    assert "NVIDIA CONFIDENTIAL" in boilerplate
    assert "Only page one" not in boilerplate
    assert "Unique" not in boilerplate
    stripped = _strip_boilerplate(
        "NVIDIA CONFIDENTIAL\nOnly page one",
        boilerplate,
    )
    assert stripped == "Only page one"


def test_page_numbered_running_header_is_detected_and_stripped() -> None:
    pages = [
        [f"Guide DG-12084-001_v1.3 | {page}", f"Unique content {page}"]
        for page in range(35, 39)
    ]

    boilerplate = detect_boilerplate(pages)

    assert "Guide DG-12084-001_v1.3" in boilerplate
    assert _strip_boilerplate("Guide DG-12084-001_v1.3 | 35\nUseful text", boilerplate) == (
        "Useful text"
    )


def test_numbered_figure_captions_are_not_boilerplate() -> None:
    pages = [[f"Figure {page} Camera routing"] for page in range(1, 5)]

    boilerplate = detect_boilerplate(pages)

    assert not boilerplate
    assert _strip_boilerplate("Figure 1 Camera routing", boilerplate) == (
        "Figure 1 Camera routing"
    )


def test_heading_rejects_nvidia_document_identifier() -> None:
    assert not _is_heading("Jetson Thor Design Guide DG-12084-001_v1.3 | 35")


def test_short_prose_and_dot_leaders_are_dropped() -> None:
    chunks = []

    _emit_prose_chunks(chunks, ["short"], ["Section"], 1)
    _emit_prose_chunks(
        chunks,
        ["List of Figures ........................................ 129"],
        ["Contents"],
        2,
    )
    useful = "A" * _MIN_PROSE_CHARS
    _emit_prose_chunks(chunks, [useful], ["Useful"], 3)

    assert [chunk.content for chunk in chunks] == [useful]


def test_table_renderer_drops_empty_grid_positions() -> None:
    assert _render_table([[None, "Category", None, None, "Function", None]]) == (
        "Category | Function"
    )


def test_table_renderer_collapses_cell_newlines() -> None:
    assert _render_table([["USB 2.0\nUSB 3.", "Display"]]) == "USB 2.0 USB 3. | Display"


def test_table_renderer_skips_fully_empty_rows() -> None:
    assert _render_table([[None, "  ", None], ["Useful", None]]) == "Useful"


def test_table_renderer_aligns_sparse_detected_grid() -> None:
    rendered = _render_table(
        [
            [None, "Category", None, None, "Function", None, None, "Category", None, None, "Function", None],
            ["USB", None, None, "USB 2.0\nUSB 3.", None, None, "Display", None, None, "HDMI / DP", None, None],
        ]
    )

    assert rendered == (
        "Category | Function | Category | Function\n"
        "USB | USB 2.0 USB 3. | Display | HDMI / DP"
    )
    assert [len(line.split(" | ")) for line in rendered.splitlines()] == [4, 4]


def test_pdf_ingest_suppresses_low_content_table(tmp_path: Path, monkeypatch) -> None:
    import fitz

    class FakeTable:
        bbox = (0, 0, 10, 10)

        def extract(self):
            content = "x" * (_MIN_TABLE_CHARS - 1)
            return [[None, content, None]]

    class FakePage:
        def get_text(self, kind, *, sort):
            assert sort is True
            return "" if kind == "text" else []

        def find_tables(self):
            return type("Finder", (), {"tables": [FakeTable()]})()

        def get_images(self, *, full):
            assert full is True
            return []

    class FakeDocument:
        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, traceback):
            return False

        def __iter__(self):
            return iter([FakePage()])

        def __len__(self):
            return 1

    monkeypatch.setattr(fitz, "open", lambda path: FakeDocument())
    source = ManifestSource(
        path=tmp_path / "guide.pdf",
        source_path="RAG_DOCS/guide.pdf",
        doc_key="guide-v1",
        title="Guide",
        doc_type="design_guide",
        platform="Thor",
        version="1.0",
    )

    records = ingest_pdf(source)

    assert records.chunks == []


def test_figure_ref_requires_a_caption() -> None:
    class FakePage:
        def get_images(self, *, full):
            assert full is True
            return [(1,)]

        def get_image_info(self):
            return []

    output = []
    page = FakePage()

    _emit_figure_ref(output, page, [], [], 1)
    assert output == []

    _emit_figure_ref(output, page, [(object(), "Figure 1 Camera routing")], [], 1)
    assert len(output) == 1
    assert output[0].chunk_type == "figure_ref"
    assert "Figure 1 Camera routing" in output[0].content


def test_prose_chunker_respects_budget_and_carries_paragraph_overlap() -> None:
    paragraphs = [f"paragraph-{index} " + (str(index) * 30) for index in range(1, 6)]

    chunks = chunk_prose(paragraphs, budget_chars=100, overlap=0.2)

    assert len(chunks) > 1
    assert all(len(chunk) <= 100 for chunk in chunks)
    first_paragraphs = chunks[0].split("\n\n")
    second_paragraphs = chunks[1].split("\n\n")
    assert first_paragraphs[-1] == second_paragraphs[0]


def test_manifest_rejects_missing_required_field(tmp_path: Path) -> None:
    manifest = tmp_path / "manifest.yaml"
    manifest.write_text(
        yaml.safe_dump(
            {
                "sources": [
                    {
                        "path": str(tmp_path / "missing.pdf"),
                        "doc_key": "missing-title",
                        "type": "design_guide",
                        "platform": "Thor",
                        "version": "1.0",
                    }
                ]
            }
        ),
        encoding="utf-8",
    )

    with pytest.raises(KnowledgeDBError, match="missing required field.*title"):
        load_manifest(manifest)


def test_manifest_rejects_missing_source_file(tmp_path: Path) -> None:
    manifest = tmp_path / "manifest.yaml"
    manifest.write_text(
        yaml.safe_dump(
            {
                "sources": [
                    {
                        "path": str(tmp_path / "missing.pdf"),
                        "doc_key": "missing-file",
                        "title": "Missing",
                        "type": "design_guide",
                        "platform": "Thor",
                        "version": "1.0",
                    }
                ]
            }
        ),
        encoding="utf-8",
    )

    with pytest.raises(KnowledgeDBError, match="document source file not found"):
        load_manifest(manifest)


def test_content_hash_is_stable_and_changes_with_file(tmp_path: Path) -> None:
    source = tmp_path / "source.bin"
    source.write_bytes(b"same bytes")

    first = _content_hash(source)
    second = _content_hash(source)
    source.write_bytes(b"different bytes")

    assert first == second
    assert len(first) == 64
    assert _content_hash(source) != first


def test_document_title_is_inserted_with_source(tmp_path: Path) -> None:
    class FakeConnection:
        def __init__(self) -> None:
            self.calls = []

        def execute(self, query, params=None):
            self.calls.append((query, params))
            return self

    source = ManifestSource(
        path=tmp_path / "guide.pdf",
        source_path="RAG_DOCS/guide.pdf",
        doc_key="guide-v1",
        title="Jetson Thor Design Guide",
        doc_type="design_guide",
        platform="Thor",
        version="1.0",
    )
    connection = FakeConnection()

    _replace_source(connection, source, "abc123", IngestedRecords(page_count=10))

    insert = next(call for call in connection.calls if "insert into doc_sources" in call[0])
    assert "title" in insert[0]
    assert insert[1]["title"] == "Jetson Thor Design Guide"
